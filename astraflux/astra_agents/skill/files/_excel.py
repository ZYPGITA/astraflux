# -*- coding: utf-8 -*-

import json
import openpyxl


class Write:
    """Write Excel files (.xlsx). Content must be JSON array of objects."""

    @staticmethod
    def write(content: str, filepath: str, encoding: str = "utf-8", sheet_name: str = "Sheet1") -> str:

        data = json.loads(content)
        if not isinstance(data, list):
            return "Data must be a JSON array of objects."
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        if data:
            # Header row from keys of first object
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h, "") for h in headers])
        wb.save(filepath)
        return f"Successfully saved {len(data)} rows to Excel at '{filepath}'."


class Read:
    """Read Excel files (.xlsx). Returns JSON array of objects."""

    @staticmethod
    def read(filepath: str, sheet_name: str = None) -> str:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return "[]"
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = val
            data.append(row_dict)
        wb.close()
        return json.dumps(data, ensure_ascii=False, indent=2)
